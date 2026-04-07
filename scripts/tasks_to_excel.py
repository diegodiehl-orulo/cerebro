#!/usr/bin/env python3
"""
tasks_to_excel.py — Gera planilha Excel das tarefas tl;dv
========================================================

Lê tasks/consolidated/all_tasks.json e gera um Excel organizado com:
  - Aba Tarefas (todas)
  - Aba Críticas (HIGH)
  - Aba Sem Dono
  - Aba Sem Prazo
  - Aba Por Responsável

Uso:
  python3 scripts/tasks_to_excel.py                    # gera para todas
  python3 scripts/tasks_to_excel.py --since YYYY-MM-DD
  python3 scripts/tasks_to_excel.py --owner "Diego Diehl"
  python3 scripts/tasks_to_excel.py --output /caminho/saida.xlsx
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERRO: openpyxl não disponível. Instale: pip install openpyxl")
    sys.exit(1)


WORKSPACE = Path("/root/.openclaw/workspace")
SOURCE_FILE = WORKSPACE / "tasks/consolidated/all_tasks.json"


# ── Estilos ────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")      # azul escuro
CRITICAL_FILL = PatternFill("solid", fgColor="FF4B4B")    # vermelho
HIGH_FILL = PatternFill("solid", fgColor="FFD93D")        # amarelo
MEDIUM_FILL = PatternFill("solid", fgColor="6BCB77")     # verde
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1E3A5F")
NORMAL_FONT = Font(name="Calibri", size=10)
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")

# Prioridade → cor
PRIORITY_FILLS = {
    "HIGH": HIGH_FILL,
    "MEDIUM": MEDIUM_FILL,
    "LOW": PatternFill("solid", fgColor="D9D9D9"),
}
# Status → cor
STATUS_FILLS = {
    "OPEN": PatternFill("solid", fgColor="FFF2CC"),
    "IN_PROGRESS": PatternFill("solid", fgColor="DAEEF3"),
    "DONE": PatternFill("solid", fgColor="E2EFDA"),
    "BLOCKED": PatternFill("solid", fgColor="FCE4D6"),
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def load_tasks(since: str = "") -> list[dict]:
    if not SOURCE_FILE.exists():
        print(f"ERRO: {SOURCE_FILE} não encontrado. Rode task_generator.py primeiro.")
        sys.exit(1)
    with open(SOURCE_FILE) as f:
        data = json.load(f)
    tasks = data.get("tasks", [])
    if since:
        tasks = [t for t in tasks if (t.get("source_date") or "") >= since]
    return tasks


def priority_order(p: str) -> int:
    return {"HIGH": 0, "MEDIUM": 1, "LOW": 2}.get(p, 3)


def apply_header(ws, headers: list[str], row: int = 1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def set_col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


def write_row(ws, row: int, values: list, priority: str = ""):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = NORMAL_FONT
        cell.alignment = WRAP
        cell.border = BORDER
        if priority and priority in PRIORITY_FILLS:
            cell.fill = PRIORITY_FILLS[priority]


# ── Abas ─────────────────────────────────────────────────────────────────────

def create_tasks_sheet(wb, tasks: list[dict]):
    ws = wb.active
    ws.title = "Tarefas"
    headers = ["ID", "Título", "Descrição", "Dono", "Prazo", "Prioridade",
               "Área", "Tipo", "Status", "Fonte", "Data Reunião",
               "Critério de Conclusão", "Bloqueio", "Criado em"]
    apply_header(ws, headers, row=1)

    # Sort: HIGH first, then by date
    tasks_sorted = sorted(tasks, key=lambda t: (priority_order(t.get("priority","LOW")), t.get("source_date","")))

    for row, t in enumerate(tasks_sorted, 2):
        priority = t.get("priority", "MEDIUM")
        status = t.get("status", "OPEN")
        write_row(ws, row, [
            t.get("id",""),
            t.get("title",""),
            t.get("description",""),
            t.get("owner","") or "",
            t.get("deadline","") or "",
            priority,
            t.get("area",""),
            t.get("type",""),
            status,
            t.get("source_meeting",""),
            t.get("source_date",""),
            t.get("completion_criteria","") or "",
            t.get("blocking_factor","") or "",
            t.get("created_at","")[:10],
        ], priority=priority)

        if status in STATUS_FILLS:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = STATUS_FILLS[status]

    # Col widths
    set_col_width(ws, 1, 18)   # ID
    set_col_width(ws, 2, 40)   # Título
    set_col_width(ws, 3, 60)   # Descrição
    set_col_width(ws, 4, 20)   # Dono
    set_col_width(ws, 5, 12)   # Prazo
    set_col_width(ws, 6, 12)   # Prioridade
    set_col_width(ws, 7, 14)   # Área
    set_col_width(ws, 8, 16)   # Tipo
    set_col_width(ws, 9, 12)   # Status
    set_col_width(ws, 10, 20)  # Fonte
    set_col_width(ws, 11, 13)  # Data
    set_col_width(ws, 12, 40)  # Critério
    set_col_width(ws, 13, 40)  # Bloqueio
    set_col_width(ws, 14, 12)  # Criado

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def create_filtered_sheet(wb, name: str, tasks: list[dict], cols: list[str], title: str):
    ws = wb.create_sheet(title=name)
    if not tasks:
        ws.cell(row=1, column=1, value="Nenhuma tarefa nesta categoria.")
        return
    headers = [c for c,_ in cols]
    apply_header(ws, headers)
    for row, t in enumerate(tasks, 2):
        for col, (header, key) in enumerate(cols, 1):
            cell = ws.cell(row=row, column=col, value=t.get(key,""))
            cell.font = NORMAL_FONT
            cell.alignment = WRAP
            cell.border = BORDER
            if key == "priority" and t.get(key) in PRIORITY_FILLS:
                cell.fill = PRIORITY_FILLS[t.get(key)]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def create_summary_sheet(wb, tasks: list[dict]):
    ws = wb.create_sheet(title="Resumo")
    # Title
    ws.cell(row=1, column=1, value="Resumo de Tarefas — tl;dv")
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    ws.cell(row=3, column=1, value=f"Total de tarefas: {len(tasks)}")
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")

    row = 5
    # By priority
    ws.cell(row=row, column=1, value="Por Prioridade")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    for p in ["HIGH", "MEDIUM", "LOW"]:
        cnt = len([t for t in tasks if t.get("priority") == p])
        ws.cell(row=row, column=1, value=p)
        ws.cell(row=row, column=2, value=cnt)
        if p in PRIORITY_FILLS:
            ws.cell(row=row, column=1).fill = PRIORITY_FILLS[p]
        row += 1

    row += 1
    # By type
    ws.cell(row=row, column=1, value="Por Tipo")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    types = {}
    for t in tasks:
        types[t.get("type","")] = types.get(t.get("type",""), 0) + 1
    for t, cnt in sorted(types.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=t)
        ws.cell(row=row, column=2, value=cnt)
        row += 1

    row += 1
    # Without owner / deadline
    ws.cell(row=row, column=1, value="Alertas")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    no_owner = len([t for t in tasks if not t.get("owner")])
    no_deadline = len([t for t in tasks if not t.get("deadline")])
    no_both = len([t for t in tasks if not t.get("owner") and not t.get("deadline")])
    ws.cell(row=row, column=1, value="Sem dono")
    ws.cell(row=row, column=2, value=no_owner)
    row += 1
    ws.cell(row=row, column=1, value="Sem prazo")
    ws.cell(row=row, column=2, value=no_deadline)
    row += 1
    ws.cell(row=row, column=1, value="Sem dono E sem prazo")
    ws.cell(row=row, column=2, value=no_both)

    set_col_width(ws, 1, 25)
    set_col_width(ws, 2, 12)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Gera Excel de tarefas tl;dv")
    parser.add_argument("--since", type=str, default="", help="YYYY-MM-DD")
    parser.add_argument("--owner", type=str, default="", help="Filtrar por responsável")
    parser.add_argument("--priority", type=str, default="", help="Filtrar por prioridade: HIGH/MEDIUM/LOW")
    parser.add_argument("--output", "-o", type=str, default="",
                        help="Caminho do arquivo de saída (default: tasks/tasks_tldv_YYYYMMDD_HHMMSS.xlsx)")
    args = parser.parse_args()

    tasks = load_tasks(since=args.since)
    if args.owner:
        tasks = [t for t in tasks if args.owner.lower() in (t.get("owner") or "").lower()]
    if args.priority:
        tasks = [t for t in tasks if t.get("priority") == args.priority.upper()]

    if not tasks:
        print("Nenhuma tarefa encontrada.")
        return

    wb = openpyxl.Workbook()

    # Resumo
    create_summary_sheet(wb, tasks)

    # Tarefas
    create_tasks_sheet(wb, tasks)

    # Críticas
    critical = [t for t in tasks if t.get("priority") == "HIGH"]
    cols = [("ID","id"),("Título","title"),("Dono","owner"),("Prazo","deadline"),("Área","area"),("Tipo","type"),("Status","status"),("Bloqueio","blocking_factor")]
    create_filtered_sheet(wb, "Criticas", critical, cols, "🔴 Críticas")

    # Sem dono
    no_owner = [t for t in tasks if not t.get("owner")]
    cols = [("ID","id"),("Título","title"),("Prazo","deadline"),("Prioridade","priority"),("Área","area"),("Tipo","type")]
    create_filtered_sheet(wb, "SemDono", no_owner, cols, "⚠️ Sem Dono")

    # Sem prazo
    no_deadline = [t for t in tasks if not t.get("deadline")]
    cols = [("ID","id"),("Título","title"),("Dono","owner"),("Prioridade","priority"),("Área","area"),("Tipo","type")]
    create_filtered_sheet(wb, "SemPrazo", no_deadline, cols, "⚠️ Sem Prazo")

    # Por responsável
    by_owner = {}
    for t in tasks:
        owner = t.get("owner") or "_SEM_DONO_"
        by_owner.setdefault(owner, []).append(t)
    for owner, owner_tasks in sorted(by_owner.items(), key=lambda x: -len(x[1])):
        safe_name = owner[:25].replace("/","_").replace("\\","_")
        cols = [("ID","id"),("Título","title"),("Prazo","deadline"),("Prioridade","priority"),("Área","area"),("Tipo","type"),("Status","status")]
        create_filtered_sheet(wb, safe_name, owner_tasks, cols, owner)

    # Save
    if not args.output:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        args.output = WORKSPACE / f"tasks/tasks_tldv_{ts}.xlsx"

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Excel gerado: {output_path}")
    print(f"Total: {len(tasks)} tarefas | {len(by_owner)} responsáveis | {len(critical)} críticas")


if __name__ == "__main__":
    main()
